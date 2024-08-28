import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_products_price_per_supplier = {}
products_under_ten = {}

for row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(row, 4).value
    inventory = product_list.cell(row, 2).value
    price = product_list.cell(row, 3).value
    inventory_price = product_list.cell(row, 5)

    # calculating products count per supplier
    if supplier_name not in products_per_supplier:
        products_per_supplier[supplier_name] = 0
    products_per_supplier[supplier_name] += 1

    # calculating total products price per supplier
    if supplier_name not in total_products_price_per_supplier:
        total_products_price_per_supplier[supplier_name] = 0
    total_products_price_per_supplier[supplier_name] += inventory * price

    # get products under 10
    if inventory < 10:
        product_number = int(product_list.cell(row, 1).value)
        products_under_ten[product_number] = int(inventory)

    # add total inventory price
    inventory_price.value = inventory * price

print(products_per_supplier)
print(total_products_price_per_supplier)
print(products_under_ten)

inv_file.save("updated_inventory.xlsx")

